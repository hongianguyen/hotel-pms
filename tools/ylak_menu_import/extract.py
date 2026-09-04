# -*- coding: utf-8 -*-
"""Extract Y Lak restaurant menu + recipes from COST.xlsx into menu_data.json.

Recipes come from the sheet 'ĐỊNH LƯỢNG NHẬP MÁY PHẦN MỀM' (the portion spec the
owner pointed at). Ingredient standard costs come from the 'A La Carte' and
'SET MENU' costing sheets. Selling prices come from the FOOD MENU deck via
deck_prices.py.

Run:  python3 extract.py
Out:  menu_data.json  +  a report on stdout
"""
import json
import re
import unicodedata
from collections import defaultdict, Counter
from pathlib import Path

import openpyxl

from deck_prices import DECK_PRICES

HERE = Path(__file__).parent
XLSX = HERE / "COST.xlsx"
OUT = HERE / "menu_data.json"

RECIPE_SHEET = "ĐỊNH LƯỢNG NHẬP MÁY PHẦN MỀM"

# Overhead line in the costing sheets - not a real ingredient, never a BoM line.
OVERHEAD_MARKER = "chi phi khac"

# ---------------------------------------------------------------------------
# Corrections applied to the source sheet. Each is reported, not silent.
# Keyed by (dish_name_fragment, ingredient_name) -> {field: new_value}
# ---------------------------------------------------------------------------
CORRECTIONS = [
    {
        "dish": "Canh chua cá lăng",
        "ingredient": "Cá Lăng",
        "field": "unit",
        "from": "trái",
        "to": "kg",
        "why": "Unit row-shift in rows 111-113: every other occurrence of Cá Lăng "
               "is kg (rows 33, 51, 114). 'trái' (fruit) is impossible for fish.",
    },
    {
        "dish": "Canh chua cá lăng",
        "ingredient": "Thơm",
        "field": "unit",
        "from": "kg",
        "to": "trái",
        "why": "Same row-shift: Thơm (pineapple) is 'trái' in rows 102, 115, 122.",
    },
    {
        "dish": "Canh chua chả cá",
        "ingredient": "Măng chua",
        "field": "unit",
        "from": "trái",
        "to": "kg",
        "why": "Măng chua is kg everywhere else (rows 113, 117, 124, 130).",
    },
    {
        "dish": "Gỏi gà hoa chuối",
        "ingredient": None,
        "field": "portion_unit",
        "from": None,
        "to": "phần",
        "why": "Row 11 has no portion unit; every comparable salad uses 'phần'.",
    },
    {
        "dish": "Club sandwich gà",
        "ingredient": "bánh mỳ san wich",
        "field": "qty",
        "from": "3 lát",
        "to": 3.0,
        "why": "Quantity written as text '3 lát' (3 slices). Taken as 3 with unit "
               "changed from 'cây' (whole loaf) to 'lát' (slice).",
    },
    {
        "dish": "Club sandwich gà",
        "ingredient": "bánh mỳ san wich",
        "field": "unit",
        "from": "cây",
        "to": "lát",
        "why": "See above - the quantity 3 refers to slices, not whole loaves.",
    },
    {
        "dish": "Picnic",
        "ingredient": "Trái cây theo mùa",
        "field": "unit",
        "from": None,
        "to": "phần",
        "why": "No unit given. 'Trái cây theo mùa' is itself a menu dish sold by "
               "'phần', so the Picnic consumes 0.5 portion of it (nested BoM).",
    },
]

# Explicit recipe-sheet name -> FOOD MENU deck name. The two documents spell the
# same dish differently (SÚP vs SOUP, Mỳ vs Mì, short kitchen name vs full menu
# name), so these are stated one by one rather than fuzzy-matched.
ALIASES = {
    "Súp bí đỏ": "SOUP BÍ ĐỎ",
    "Salad rau xanh": "SALAD RAU XANH",
    "Gỏi cuốn vườn Lak": "GỎI CUỐN TÔM THỊT VƯỜN LẮK",
    "Chả Giò": "CHẢ GIÒ TÔM THỊT",
    "Cá rô phi chiên ăn kèm mắm xoài": "CÁ RÔ PHI CHIÊN GIÒN CHẤM MẮM XOÀI",
    "Cá diêu hồng sốt chanh dây": "CÁ DIÊU HỒNG CHIÊN SỐT CHANH DÂY",
    "Tép Hồ Lak ram": "TÉP HỒ LAK RAM MẶN VỚI KHẾ",
    "Chả cá hồ Lak": "CHẢ CÁ HỒ LĂK",
    "Cá lăng ướp nghệ nướng": "CÁ LĂNG ƯỚP NGHỆ NƯỚNG LÁ CHUỐI",
    "Sườn nường BBQ": "SƯỜN NƯỚNG BBQ CƠM LAM",
    "Gà nướng": "GÀ NƯỚNG DÙNG KÈM CƠM LAM",
    "Vịt quay": "VỊT QUAY HỒ LĂK SỐT TIÊU ĐEN",
    "Heo nướng ống tre": "HEO NƯỚNG ỐNG TRE ĂN KÈM CƠM LAM",
    "Mỳ Ý sốt cà chua": "MÌ Ý SỐT CÀ CHUA",
    "Mỳ ý sốt bò bằm": "MÌ Ý SỐT BÒ BẰM",
    "Mì xào rau củ": "MÌ XÀO RAU CỦ",
    "Mì xào bò": "MÌ XÀO BÒ",
    "Miến xào nồi đất": "MIẾN XÀO",
    "Cơm chiên rau củ hạt sen": "CƠM CHIÊN RAU CỦ HẠT SEN",
    "Cơm chiên gà xé hạt sen": "CƠM CHIÊN GÀ HẠT SEN",
    "Cơm chiên hải sản hạt sen": "CƠM CHIÊN HẢI SẢN HẠT SEN",
    "Bò lúc lắc + Khoai tây chiên": "BÒ LÚC LẮC KHOAI TÂY CHIÊN",
    "Bò beefsteak + Khoai tây chiên": "BÒ BEEFSTEAK KHOAI TÂY CHIÊN",
    "Canh tập tàng tôm": "CANH TÔM TẬP TÀNG",
    "Canh cua tập tàng": "CANH CUA TẬP TÀNG",
    "Canh chua chả cá": "CANH CHUA CHẢ CÁ THÁC LÁC",
    "Lẩu cá lăng đồng": "LẨU CÁ LĂNG ĐỒNG NẤU MĂNG CHUA",
    "Lẩu cá thác lác": "LẨU CHẢ CÁ THÁC LÁC MĂNG CHUA",
    "Bánh chuối nướng +kem": "BÁNH CHUỐI NƯỚNG DÙNG KÈM KEM VANI",
    "Sữa chua": "SỮA CHUA TRÁI CÂY THEO MÙA",
    "Kem socola": "KEM SÔ CÔ LA",
    "Kem vanilla": "KEM VANILLA",
    "Trứng chiên": "TRỨNG VỊT HỒ LAK CHIÊN HÀNH",
    "Khoai tây chiên": "KHOAI TÂY CHIÊN",
    "Đậu hủ sốt cà chua": "ĐẬU HỦ SỐT CÀ CHUA",
    "Đậu hủ sốt sả": "ĐẬU HỦ SỐT SẢ",
    "Cà tím kho tộ": "CÀ TÍM KHO TỘ",
    "Cà tím nướng mỡ hành": "CÀ TÍM NƯỚNG MỠ HÀNH",
    "Cá bông lau kho tộ": "CÁ BÔNG LAU KHO TỘ",
    "Rau muống xào tỏi": "RAU MUỐNG XÀO TỎI",
    "Rau rừng xào tỏi": "RAU RỪNG XÀO TỎI",
    "Rau rừng luộc kho quẹt": "RAU RỪNG LUỘC KHO QUẸT",
    "Lá bép xào tỏi": "LÁ BÉP XÀO TỎI",
    "Lá bép xào cá hộp": "LÁ BÉP XÀO CÁ HỘP",
    "Gỏi rau muống tép ram": "GỎI RAU MUỐNG TÉP RAM",
}

# Dishes that are prep/semi-finished items, not sold on their own.
# These get a NORMAL bom (made in batches), everything else gets a KIT bom.
PREP_ITEMS = {"ỐNG CƠM LAM"}

# Dishes in the recipe sheet that are packages/room-service, not a-la-carte menu
# items. Kept as products but flagged so the user can decide where they live.
NON_MENU_DISHES = {"Set up Bungalow", "Picnic ( cho 1 pax)"}


def strip_accents(s):
    s = unicodedata.normalize("NFD", str(s))
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


def norm(s):
    """Normalise a Vietnamese name for matching."""
    if s is None:
        return ""
    s = strip_accents(s).lower()
    s = re.sub(r"\(.*?\)", " ", s)          # drop parentheticals
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return " ".join(s.split())


def vn_part(name):
    """'Súp bí đỏ/ Pumpkin soup' -> 'Súp bí đỏ'."""
    return name.split("/")[0].strip()


def en_part(name):
    parts = name.split("/", 1)
    return parts[1].strip() if len(parts) > 1 else ""


# ---------------------------------------------------------------------------
# 1. Recipes
# ---------------------------------------------------------------------------
def read_recipes(wb):
    ws = wb[RECIPE_SHEET]
    dishes, cur = [], None
    for r in range(2, ws.max_row + 1):
        a, b, c, d, e = [ws.cell(r, i).value for i in range(1, 6)]
        if a and str(a).strip():
            cur = {
                "name": str(a).strip(),
                "name_vn": vn_part(str(a).strip()),
                "name_en": en_part(str(a).strip()),
                "portion_unit": str(b).strip() if b else None,
                "row": r,
                "components": [],
            }
            dishes.append(cur)
        if c and str(c).strip():
            if cur is None:
                raise RuntimeError(f"ingredient row {r} before any dish")
            cur["components"].append({
                "ingredient": str(c).strip(),
                "unit": str(d).strip() if d else None,
                "qty": e,
                "row": r,
            })
    return dishes


def apply_corrections(dishes):
    applied, unmatched = [], []
    for fix in CORRECTIONS:
        hit = False
        for dish in dishes:
            if norm(fix["dish"]) not in norm(dish["name"]):
                continue
            if fix["ingredient"] is None:
                if dish.get(fix["field"]) == fix["from"]:
                    dish[fix["field"]] = fix["to"]
                    applied.append((dish["name"], "-", fix))
                    hit = True
            else:
                for comp in dish["components"]:
                    if norm(comp["ingredient"]) != norm(fix["ingredient"]):
                        continue
                    if comp.get(fix["field"]) == fix["from"]:
                        comp[fix["field"]] = fix["to"]
                        applied.append((dish["name"], comp["ingredient"], fix))
                        hit = True
        if not hit:
            unmatched.append(fix)
    return applied, unmatched


# ---------------------------------------------------------------------------
# 2. Ingredient standard costs from the costing sheets
# ---------------------------------------------------------------------------
def read_costs(wb):
    """Collect every (ingredient, unit, unit_price) seen in the costing sheets."""
    seen = defaultdict(list)

    def harvest(ws, name_col, unit_col, price_col):
        for r in range(1, ws.max_row + 1):
            n = ws.cell(r, name_col).value
            u = ws.cell(r, unit_col).value
            p = ws.cell(r, price_col).value
            if not n or not isinstance(p, (int, float)) or p <= 0:
                continue
            n = str(n).strip()
            if OVERHEAD_MARKER in norm(n) or norm(n) in ("tong", "tong cong"):
                continue
            seen[norm(n)].append({
                "raw_name": n,
                "unit": str(u).strip() if u else None,
                "price": float(p),
            })

    harvest(wb["A La Carte"], 1, 2, 3)
    harvest(wb["SET MENU"], 1, 2, 3)   # left-hand table
    harvest(wb["SET MENU"], 7, 8, 9)   # right-hand table
    return seen


def pick_cost(entries):
    """Most common unit; highest price seen for that unit (most recent buy)."""
    units = Counter(e["unit"] for e in entries if e["unit"])
    unit = units.most_common(1)[0][0] if units else None
    prices = [e["price"] for e in entries if e["unit"] == unit] or \
             [e["price"] for e in entries]
    return unit, max(prices)


# ---------------------------------------------------------------------------
# 3. Selling prices
# ---------------------------------------------------------------------------
def match_prices(dishes):
    index = {}
    for vn, en, price, section in DECK_PRICES:
        index[norm(vn)] = (vn, en, price, section)

    alias = {norm(k): norm(v) for k, v in ALIASES.items()}

    matched, unmatched, via_alias = [], [], []
    for dish in dishes:
        key = norm(dish["name_vn"])
        hit = index.get(key)
        if hit is None and key in alias:
            hit = index.get(alias[key])
            if hit:
                via_alias.append((dish["name"], hit[0]))
        if hit:
            dish["sale_price"] = hit[2]
            dish["menu_section"] = hit[3]
            dish["deck_name"] = hit[0]
            dish["name_en_deck"] = hit[1]
            matched.append((dish["name"], hit[0], hit[2]))
        else:
            dish["sale_price"] = None
            dish["menu_section"] = "Chưa phân loại"
            dish["deck_name"] = None
            unmatched.append(dish["name"])
    return matched, unmatched, via_alias


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    dishes = read_recipes(wb)
    applied, bad_fixes = apply_corrections(dishes)
    costs = read_costs(wb)
    matched, unmatched, via_alias = match_prices(dishes)

    # ingredient master
    ing_units = defaultdict(set)
    for d in dishes:
        for c in d["components"]:
            ing_units[c["ingredient"]].add(c["unit"])

    ingredients = {}
    no_cost = []
    for name, units in sorted(ing_units.items()):
        entries = costs.get(norm(name))
        if entries:
            _, price = pick_cost(entries)
        else:
            price = 0.0
            no_cost.append(name)
        ingredients[name] = {
            "name": name,
            "units_used": sorted(u for u in units if u),
            "uom": sorted(u for u in units if u)[0] if any(units) else "kg",
            "standard_cost": round(price, 2),
        }

    # nested items: a component that is itself a dish
    dish_by_norm = {norm(d["name_vn"]): d["name"] for d in dishes}
    nested = []
    for d in dishes:
        for c in d["components"]:
            t = dish_by_norm.get(norm(c["ingredient"]))
            if t and t != d["name"]:
                nested.append((d["name"], c["ingredient"], t))
    for p in PREP_ITEMS:
        if p in ing_units:
            nested.append(("(prep item)", p, p))

    data = {
        "dishes": dishes,
        "ingredients": ingredients,
        "prep_items": sorted(PREP_ITEMS),
        "non_menu_dishes": sorted(NON_MENU_DISHES),
        "nested": nested,
    }
    OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2))

    # ---------------- report ----------------
    print("=" * 74)
    print("Y LAK MENU EXTRACTION REPORT")
    print("=" * 74)
    print(f"Dishes (recipes)      : {len(dishes)}")
    print(f"Component lines       : {sum(len(d['components']) for d in dishes)}")
    print(f"Distinct ingredients  : {len(ingredients)}")
    print(f"Priced from deck      : {len(matched)}")
    print(f"NO price match        : {len(unmatched)}")
    print(f"Ingredients w/o cost  : {len(no_cost)}")

    print("\n--- CORRECTIONS APPLIED TO SOURCE DATA ---")
    for dish, ing, fix in applied:
        print(f"  [{dish}] {ing}: {fix['field']} {fix['from']!r} -> {fix['to']!r}")
        print(f"      why: {fix['why']}")
    if bad_fixes:
        print("  !! correction did not match any row (source may have changed):")
        for f in bad_fixes:
            print(f"     {f['dish']} / {f['ingredient']} / {f['field']}")

    print(f"\n--- MATCHED VIA EXPLICIT ALIAS ({len(via_alias)}) ---")
    for a, b in via_alias:
        print(f"   {a}  ->  {b}")

    print("\n--- DISHES WITH NO SELLING PRICE (need owner input) ---")
    for n in unmatched:
        print("  ", n)

    print("\n--- INGREDIENTS WITH NO COST IN THE COSTING SHEETS ---")
    for n in no_cost:
        print("  ", n, "  units:", ingredients[n]["units_used"])

    print("\n--- NESTED ITEMS (a component that is itself a dish/prep) ---")
    for parent, comp, target in nested:
        print(f"   {parent}  ->  {comp}")

    print(f"\nWrote {OUT}")


if __name__ == "__main__":
    main()
